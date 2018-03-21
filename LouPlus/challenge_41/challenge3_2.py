# -*- coding:utf8 -*-
from openpyxl import load_workbook #可以用来载入已有数据表格
from openpyxl import Workbook #可以用来处理新的数据表格
import datetime #可以用来处理时间相关的数据

#读取文件及表
wb = load_workbook('courses.xlsx')
StudentSheet = wb['students']
TimeSheet = wb['time']

def combine():
    """
    该函数可以用来处理原数据文件
    1.合并表格写入的combine表中
    2.保存原数据文件

    """
    # 创建 combine 表
    CombineSheet = wb.create_sheet(title='combine')
    # 添加第一行
    CombineSheet.append(['创建时间','课程名称','学习人数','学习时间'])
    # 合并表格并依次添加到 combine 表中
    for s in StudentSheet.values:
        # 去掉包含表头的一行
        if s[2] != '学习人数':
            # 遍历匹配学习时间数据
            for t in TimeSheet.values:
                if t[1] == s[1]:
                # 添加记录到 combine 表中
                    CombineSheet.append(list(s) + [t[2]])
    # 覆盖保存
    wb.save('courses.xlsx')
    


def split():
    """
    该函数可以用来分割文件
    1.读取combine表中的数据
    2.将数据按时间分割
    3.写入不同的数据表中

    """
    CombineSheet = wb['combine']
    # 存储 combine 表中的年份
    split_name = []
    # 遍历表获取每条数据对应的年份
    for item in CombineSheet.values:
        if item[0] != '创建时间':
            split_name.append(item[0].strftime("%Y"))

    # 分别储存数据
    for name in set(split_name):
        # 创建文件
        wb_temp = Workbook()
        # 删除已有的默认 Sheet 表
        wb_temp.remove(wb_temp.active)
        # 创建相应的年份命名的表
        ws = wb_temp.create_sheet(title=name)
        # 写入相应年份的数据
        for item_by_year in CombineSheet.values:
            if item_by_year[0] != '创建时间':
                if item_by_year[0].strftime("%Y") == name:
                    ws.append(item_by_year)
        # 存储相应年份的数据文件
        wb_temp.save(f'{name}.xlsx')

if __name__ == "__main__":
    combine()
    split()

